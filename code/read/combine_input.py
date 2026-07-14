"""
Compare sheet names between two Excel files.

Usage:
    python compare_sheet_names.py file1.xlsx file2.xlsx
"""
import json
import os
import sys
from openpyxl import load_workbook
os.chdir(os.path.dirname(os.path.abspath(__file__)))
os.chdir('..')
NUM_ROWS = 1  # how many rows (from the top) to show


def get_row_values(ws, row, max_col):
    return [ws.cell(row=row, column=c).value for c in range(1, max_col + 1)]


def compare_sheets(file1_path, file2_path):
    wb1 = load_workbook(file1_path, read_only=True)
    wb2 = load_workbook(file2_path, read_only=True)

    sheets1 = set(wb1.sheetnames)
    sheets2 = set(wb2.sheetnames)

    common = sorted(sheets1 & sheets2)
    only1 = sorted(sheets1 - sheets2)
    only2 = sorted(sheets2 - sheets1)

    print(f"File 1 ({file1_path}): {len(sheets1)} sheets")
    print(f"File 2 ({file2_path}): {len(sheets2)} sheets\n")

    print(f"Common sheets ({len(common)}):")
    for s in common:
        cols1 = wb1[s].max_column
        cols2 = wb2[s].max_column
        rows1 = wb1[s].max_row
        rows2 = wb2[s].max_row
        match = "OK" if cols1 == cols2 else "DIFFERS"
        print(f"{s:<40}{cols1:<12}{cols2:<12}{match}")
        if match=="DIFFERS":
            for r in range(1, NUM_ROWS + 1):
                row1 = get_row_values(wb1[s], r, cols1)
                row2 = get_row_values(wb2[s], r, cols2)
                print(f"\n-- Row {r} --")
                print(f"  File1: {row1}")
                print(f"  File2: {row2}")


        match = "OK" if rows1 == rows2 else "DIFFERS"
        print(f"{s:<40}{rows1:<12}{rows2:<12}{match}")


    print(f"\nOnly in file 1 ({len(only1)}):")
    for s in only1:
        print(f"  - {s}")

    print(f"\nOnly in file 2 ({len(only2)}):")
    for s in only2:
        print(f"  - {s}")

# Define file name
json_settings_file = f'settings/IESA_settings_v.combine.json'

# Read file
with open(json_settings_file, 'r') as file:
    json_settings_str = file.read()

json_settings = json.loads(json_settings_str)

# Decompose the JSON struct
file_name_prio = json_settings['file_name_prio']
file_name_second = json_settings['file_name_second']
scenario_name = json_settings['scenario_name']
read_input = json_settings['read_input']
save_output = json_settings['save_output']
file1_path = os.path.join('input', file_name_prio)
file2_path = os.path.join('input', file_name_second)
compare_sheets(file1_path, file2_path)